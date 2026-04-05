"""Génération de fichiers Excel pour les écritures comptables (PCGM)."""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DARK_NAVY = "1E2A4A"
LIGHT_BLUE = "EFF6FF"
WHITE = "FFFFFF"
GRAY_BG = "F8FAFC"
GREEN = "16A34A"
RED = "DC2626"
BORDER_COLOR = "E5E7EB"

_thin = Side(style="thin", color=BORDER_COLOR)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_COL_WIDTHS = {
    "A": 6, "B": 16, "C": 12, "D": 12,
    "E": 40, "F": 18, "G": 18, "H": 25,
}


def generate_journal_excel(
    invoice: dict,
    accounts: list,
    client: dict,
    company: dict,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Comptable"

    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # ── Company header ────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = company.get("name", "").upper()
    c.font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    c.fill = PatternFill("solid", fgColor=DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = (
        f"ICE: {company.get('ice', 'N/A')}   "
        f"IF: {company.get('if_number', 'N/A')}   "
        f"RC: {company.get('rc', 'N/A')}   "
        f"Tél: {company.get('phone', 'N/A')}"
    )
    c.font = Font(size=9, color="6B7280", name="Calibri")
    c.fill = PatternFill("solid", fgColor=DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 30
    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value = "JOURNAL DES ACHATS" if invoice.get("direction") == "achat" else "JOURNAL DES VENTES"
    c.font = Font(bold=True, size=13, color=DARK_NAVY, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Invoice meta ──────────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 18
    ws.merge_cells("A5:D5")
    c = ws["A5"]
    c.value = f"Facture N°: {invoice.get('invoice_number', 'N/A')}"
    c.font = Font(size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=GRAY_BG)

    ws.merge_cells("E5:H5")
    c = ws["E5"]
    c.value = (
        f"Client: {client.get('name', 'N/A')}   "
        f"Fournisseur: {invoice.get('supplier_name', 'N/A')}"
    )
    c.font = Font(size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=GRAY_BG)
    c.alignment = Alignment(horizontal="right")

    ws.row_dimensions[6].height = 18
    ws.merge_cells("A6:D6")
    c = ws["A6"]
    c.value = f"Date: {invoice.get('date', 'N/A')}   TVA: {invoice.get('tva_rate', 20)}%"
    c.font = Font(size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=GRAY_BG)

    ws.merge_cells("E6:H6")
    c = ws["E6"]
    c.value = f"Exporté le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    c.font = Font(size=10, color="6B7280", name="Calibri")
    c.fill = PatternFill("solid", fgColor=GRAY_BG)
    c.alignment = Alignment(horizontal="right")

    # ── Table header ──────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 28
    headers = ["N°", "Date", "Journal", "Compte", "Libellé écriture", "Débit (MAD)", "Crédit (MAD)", "Réf. pièce"]
    cols = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for col, header in zip(cols, headers):
        c = ws[f"{col}8"]
        c.value = header
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill = PatternFill("solid", fgColor=DARK_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border

    # ── Account rows ──────────────────────────────────────────────────────────
    invoice_date = invoice.get("date", datetime.now().strftime("%d/%m/%Y"))
    journal = "ACH" if invoice.get("direction") == "achat" else "VTE"
    ref = invoice.get("invoice_number", "N/A")

    debit_accounts = [a for a in accounts if a.get("sens") == "debit"]
    credit_accounts = [a for a in accounts if a.get("sens") == "credit"]
    sorted_accounts = debit_accounts + credit_accounts

    total_debit = 0.0
    total_credit = 0.0
    row_num = 9

    for i, account in enumerate(sorted_accounts, 1):
        ws.row_dimensions[row_num].height = 22
        is_debit = account.get("sens") == "debit"
        montant = float(account.get("montant", 0))
        bg = WHITE if i % 2 == 0 else GRAY_BG

        row_data = {
            "A": i,
            "B": invoice_date,
            "C": journal,
            "D": account.get("code", ""),
            "E": account.get("libelle", ""),
            "F": montant if is_debit else "",
            "G": montant if not is_debit else "",
            "H": ref,
        }

        for col, value in row_data.items():
            c = ws[f"{col}{row_num}"]
            c.value = value
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = _border

            if col in ("F", "G") and value != "":
                c.font = Font(size=10, name="Calibri", color=GREEN if is_debit else RED)
                c.number_format = '#,##0.00 "MAD"'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif col in ("A", "B", "C", "D", "H"):
                c.font = Font(size=10, name="Calibri", color="111827")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = Font(size=10, name="Calibri", color="111827")
                c.alignment = Alignment(horizontal="left", vertical="center")

        if is_debit:
            total_debit += montant
        else:
            total_credit += montant
        row_num += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    ws.row_dimensions[row_num].height = 26
    totals = {"A": "", "B": "", "C": "", "D": "", "E": "TOTAUX", "F": total_debit, "G": total_credit, "H": ""}
    for col, value in totals.items():
        c = ws[f"{col}{row_num}"]
        c.value = value
        c.font = Font(bold=True, size=11, color=WHITE, name="Calibri")
        c.fill = PatternFill("solid", fgColor=DARK_NAVY)
        c.border = _border
        c.alignment = Alignment(
            horizontal="right" if col in ("E", "F", "G") else "center",
            vertical="center",
        )
        if col in ("F", "G"):
            c.number_format = '#,##0.00 "MAD"'

    # ── Balance check ─────────────────────────────────────────────────────────
    row_num += 1
    ws.row_dimensions[row_num].height = 20
    ws.merge_cells(f"A{row_num}:H{row_num}")
    is_balanced = abs(total_debit - total_credit) < 0.01
    c = ws[f"A{row_num}"]
    if is_balanced:
        c.value = f"Écriture équilibrée — Débit: {total_debit:,.2f} MAD = Crédit: {total_credit:,.2f} MAD"
        c.font = Font(bold=True, size=10, name="Calibri", color=GREEN)
        c.fill = PatternFill("solid", fgColor="F0FDF4")
    else:
        c.value = f"DÉSÉQUILIBRE: {abs(total_debit - total_credit):,.2f} MAD"
        c.font = Font(bold=True, size=10, name="Calibri", color=RED)
        c.fill = PatternFill("solid", fgColor="FEF2F2")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Footer ────────────────────────────────────────────────────────────────
    row_num += 2
    ws.merge_cells(f"A{row_num}:H{row_num}")
    c = ws[f"A{row_num}"]
    c.value = "Document généré par ComptaFlow — Conforme au Plan Comptable Général Marocain (PCGM)"
    c.font = Font(size=8, color="9CA3AF", italic=True, name="Calibri")
    c.alignment = Alignment(horizontal="center")

    # ── Print settings ────────────────────────────────────────────────────────
    ws.freeze_panes = "A9"
    ws.print_title_rows = "1:8"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
