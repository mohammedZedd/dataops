import io
import re
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from fastapi import status as http_status

from app.dependencies.auth import get_current_user
from app.models.client import Client
from app.models.document import Document
from app.models.invoice import Invoice
from app.models.user import User, UserRole
from app.db.dependencies import get_db
from app.schemas.document import DocumentCreate, DocumentRead, ExtractionResult, ManualInvoiceCreate, PresignedUrlResponse
from app.schemas.invoice import InvoiceRead
from app.services import client_service, document_service, invoice_service, s3_service
from app.utils.textract_service import textract_service, TextractError

router = APIRouter(tags=["documents"])

_ALLOWED_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "audio/webm",
    "audio/mp4",
    "audio/mpeg",
    "audio/ogg",
    "audio/wav",
}
_MAX_SIZE = 10 * 1024 * 1024  # 10 MB


def _slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text


def _build_s3_key(user: User, filename: str) -> str:
    """Construit la clé S3 avec noms slugifiés selon le rôle de l'utilisateur."""
    safe_name = Path(filename).name  # évite les path traversal
    company_slug = _slugify(user.company.name)
    person_slug = _slugify(f"{user.first_name} {user.last_name}")
    if user.role == UserRole.CLIENT:
        return f"{company_slug}/clients/{person_slug}/{safe_name}"
    return f"{company_slug}/accountants/{person_slug}/{safe_name}"


def _check_ownership(doc: Document, current_user: User) -> None:
    # Allow access if user uploaded the doc OR if it was sent to their client profile
    if doc.uploaded_by_user_id == current_user.id:
        return
    if current_user.client_id and doc.client_id == current_user.client_id:
        return
    raise HTTPException(status_code=403, detail="Accès refusé.")


# ─── Upload ───────────────────────────────────────────────────────────────────

_AUDIO_TYPES = {
    "audio/webm", "audio/mp4", "audio/mpeg", "audio/ogg", "audio/wav",
}


@router.post("/documents/upload", response_model=DocumentRead, status_code=201)
async def upload_document(
    file: UploadFile = File(...),
    description: str = Form(""),
    client_id: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if getattr(current_user, 'access_level', 'full') == 'readonly':
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="Votre accès est en lecture seule. Vous ne pouvez plus envoyer de documents. Contactez votre cabinet comptable.",
        )

    if file.content_type not in _ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Type non autorisé. PDF, JPG, PNG, XLSX, ou audio acceptés.")

    content = await file.read()
    if len(content) > _MAX_SIZE:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10 Mo).")

    original_name = file.filename or "file"
    s3_key = _build_s3_key(current_user, original_name)

    s3_service.upload_file(
        file_obj=io.BytesIO(content),
        s3_key=s3_key,
        content_type=file.content_type or "application/octet-stream",
    )

    # Resolve client_id
    target_client_id: str | None = None
    if client_id and current_user.role in (UserRole.ADMIN, UserRole.ACCOUNTANT):
        # Admin/accountant uploading on behalf of a client
        target_client = client_service.get_client(db, client_id)
        if not target_client or target_client.company_id != current_user.company_id:
            raise HTTPException(status_code=404, detail="Client introuvable.")
        target_client_id = client_id
    else:
        target_client_id = current_user.client_id
    if not target_client_id and current_user.role == UserRole.CLIENT:
        new_client = Client(
            name=f"{current_user.first_name} {current_user.last_name}",
            company_id=current_user.company_id,
        )
        db.add(new_client)
        db.flush()
        current_user.client_id = new_client.id
        target_client_id = new_client.id

    is_audio = file.content_type in _AUDIO_TYPES
    source_val = "cabinet" if current_user.role in (UserRole.ADMIN, UserRole.ACCOUNTANT) and target_client_id else "client"
    doc = Document(
        client_id=target_client_id,
        uploaded_by_user_id=current_user.id,
        file_name=original_name,
        s3_key=s3_key,
        source=source_val,
        file_size=len(content),
        doc_type="audio" if is_audio else None,
        description=description.strip() or None,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    # Notify admins/accountants about new upload
    if target_client_id and current_user.role == UserRole.CLIENT:
        from app.services import notification_service
        client_obj = client_service.get_client(db, target_client_id)
        notification_service.notify_staff(
            db, company_id=current_user.company_id, type="document_uploaded",
            title="Nouveau document reçu",
            message=f"{current_user.first_name} {current_user.last_name} a envoyé : {original_name}",
            link=f"/clients/{target_client_id}?tab=documents&highlight={doc.id}",
            client_id=target_client_id, document_id=doc.id,
        )
        db.commit()

    return document_service._to_read(doc)


# ─── Preview (presigned inline) ───────────────────────────────────────────────

@router.get("/documents/preview/{document_id}", response_model=PresignedUrlResponse)
def preview_document(
    document_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable.")

    if current_user.role == UserRole.CLIENT:
        _check_ownership(doc, current_user)
    else:
        if doc.client_id:
            client = client_service.get_client(db, doc.client_id)
            if not client or client.company_id != current_user.company_id:
                raise HTTPException(status_code=403, detail="Accès refusé.")
        elif doc.uploaded_by_user_id != current_user.id:
            raise HTTPException(status_code=403, detail="Accès refusé.")

    url = s3_service.generate_presigned_url_inline(doc.s3_key, doc.file_name)
    return PresignedUrlResponse(url=url)


# ─── Download (presigned attachment) ──────────────────────────────────────────

@router.get("/documents/download/{document_id}", response_model=PresignedUrlResponse)
def download_document(
    document_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable.")

    if current_user.role == UserRole.CLIENT:
        _check_ownership(doc, current_user)
    else:
        # Admins/accountants can download documents belonging to their company's clients
        if doc.client_id:
            client = client_service.get_client(db, doc.client_id)
            if not client or client.company_id != current_user.company_id:
                raise HTTPException(status_code=403, detail="Accès refusé.")
        elif doc.uploaded_by_user_id != current_user.id:
            raise HTTPException(status_code=403, detail="Accès refusé.")

    url = s3_service.generate_presigned_url_attachment(doc.s3_key, doc.file_name)
    return PresignedUrlResponse(url=url)


# ─── Delete ───────────────────────────────────────────────────────────────────

@router.delete("/documents/{document_id}", status_code=204)
def delete_document(
    document_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable.")
    _check_ownership(doc, current_user)

    s3_service.delete_file(doc.s3_key)
    db.delete(doc)
    db.commit()


# ─── List (user's own) ────────────────────────────────────────────────────────

# ─── Admin: all documents + stats ────────────────────────────────────────────

@router.get("/documents/stats")
def get_documents_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import datetime, timedelta
    from sqlalchemy import or_

    if current_user.role == UserRole.CLIENT:
        raise HTTPException(status_code=http_status.HTTP_403_FORBIDDEN, detail="Accès refusé.")

    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    three_days_ago = now - timedelta(days=3)

    from app.models.invoice import Invoice as InvoiceModel

    company_docs = (
        db.query(Document)
        .join(Client, Document.client_id == Client.id)
        .filter(Client.company_id == current_user.company_id)
    )

    total_month = company_docs.filter(Document.uploaded_at >= start_of_month).count()

    pending = (
        company_docs
        .outerjoin(InvoiceModel, InvoiceModel.document_id == Document.id)
        .filter(or_(InvoiceModel.id == None, InvoiceModel.status == "to_review"))
        .count()
    )

    validated = (
        db.query(InvoiceModel)
        .join(Document, InvoiceModel.document_id == Document.id)
        .join(Client, Document.client_id == Client.id)
        .filter(Client.company_id == current_user.company_id, InvoiceModel.status == "validated")
        .count()
    )

    rejected = (
        db.query(InvoiceModel)
        .join(Document, InvoiceModel.document_id == Document.id)
        .join(Client, Document.client_id == Client.id)
        .filter(Client.company_id == current_user.company_id, InvoiceModel.status == "rejected")
        .count()
    )

    urgent = (
        company_docs
        .outerjoin(InvoiceModel, InvoiceModel.document_id == Document.id)
        .filter(
            or_(InvoiceModel.id == None, InvoiceModel.status == "to_review"),
            Document.uploaded_at <= three_days_ago,
        )
        .count()
    )

    return {
        "total_this_month": total_month,
        "pending": pending,
        "validated": validated,
        "rejected": rejected,
        "urgent": urgent,
    }


@router.get("/documents/by-client")
def get_documents_by_client_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import datetime, timedelta
    from sqlalchemy import or_
    from app.models.invoice import Invoice as InvoiceModel

    if current_user.role == UserRole.CLIENT:
        raise HTTPException(status_code=http_status.HTTP_403_FORBIDDEN, detail="Accès refusé.")

    three_days_ago = datetime.utcnow() - timedelta(days=3)
    clients_list = db.query(Client).filter(Client.company_id == current_user.company_id).all()

    result = []
    for c in clients_list:
        docs = (
            db.query(Document)
            .filter(Document.client_id == c.id)
            .order_by(Document.uploaded_at.desc())
            .all()
        )
        if not docs:
            continue

        pending = 0
        validated = 0
        rejected = 0
        urgent = 0
        for d in docs:
            inv = d.invoice
            if inv is None or inv.status == "to_review":
                pending += 1
                if d.uploaded_at <= three_days_ago:
                    urgent += 1
            elif inv.status == "validated":
                validated += 1
            elif inv.status == "rejected":
                rejected += 1

        # Find linked user for name/email
        linked_user = db.query(User).filter(User.client_id == c.id).first()

        result.append({
            "client": {
                "id": c.id,
                "name": c.name,
                "full_name": f"{linked_user.first_name} {linked_user.last_name}" if linked_user else c.name,
                "email": linked_user.email if linked_user else None,
            },
            "total_documents": len(docs),
            "pending_count": pending,
            "validated_count": validated,
            "rejected_count": rejected,
            "urgent_count": urgent,
            "last_upload_at": docs[0].uploaded_at.isoformat(),
            "recent_files": [d.file_name for d in docs[:3]],
        })

    result.sort(key=lambda x: x["last_upload_at"], reverse=True)
    return result


@router.get("/documents/all", response_model=list[DocumentRead])
def list_all_documents(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.CLIENT:
        raise HTTPException(status_code=http_status.HTTP_403_FORBIDDEN, detail="Accès refusé.")

    docs = (
        db.query(Document)
        .join(Client, Document.client_id == Client.id)
        .filter(Client.company_id == current_user.company_id)
        .order_by(Document.uploaded_at.desc())
        .limit(200)
        .all()
    )
    return [document_service._to_read(d) for d in docs]


@router.patch("/documents/{document_id}/viewed")
def mark_document_viewed(
    document_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.CLIENT:
        raise HTTPException(status_code=http_status.HTTP_403_FORBIDDEN, detail="Accès refusé.")
    doc = document_service.get_document(db, document_id)
    if doc:
        doc.is_new = False
        db.commit()
    return {"status": "ok"}


@router.get("/documents/my", response_model=list[DocumentRead])
def list_my_documents(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Return docs uploaded by user + docs sent to their client_id by cabinet
    from sqlalchemy import or_
    docs = (
        db.query(Document)
        .filter(or_(
            Document.uploaded_by_user_id == current_user.id,
            Document.client_id == current_user.client_id,
        ) if current_user.client_id else Document.uploaded_by_user_id == current_user.id)
        .order_by(Document.uploaded_at.desc())
        .all()
    )
    return [document_service._to_read(d) for d in docs]


# ─── Existing accounting routes (admin/accountant) ────────────────────────────

@router.get("/clients/{client_id}/documents", response_model=list[DocumentRead])
def list_client_documents(
    client_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    client = client_service.get_client(db, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client introuvable.")
    if current_user.role == UserRole.CLIENT and current_user.client_id != client_id:
        raise HTTPException(status_code=403, detail="Accès refusé.")
    if current_user.role != UserRole.CLIENT and client.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Accès refusé.")
    return document_service.get_documents_by_client(db, client_id)


@router.post("/documents", response_model=DocumentRead, status_code=201)
def create_document(
    payload: DocumentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    client = client_service.get_client(db, payload.client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client introuvable.")
    if current_user.role == UserRole.CLIENT and current_user.client_id != payload.client_id:
        raise HTTPException(status_code=403, detail="Accès refusé.")
    if current_user.role != UserRole.CLIENT and client.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Accès refusé.")
    return document_service.create_document(db, payload)


@router.get("/documents/{document_id}", response_model=DocumentRead)
def get_document(
    document_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = document_service.get_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable.")
    client = client_service.get_client(db, doc.client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client introuvable.")
    if current_user.role == UserRole.CLIENT and current_user.client_id != doc.client_id:
        raise HTTPException(status_code=403, detail="Accès refusé.")
    if current_user.role != UserRole.CLIENT and client.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Accès refusé.")
    return document_service._to_read(doc)


# ─── Manual invoice creation from document ───────────────────────────────────

@router.post("/documents/{document_id}/create-invoice", response_model=InvoiceRead, status_code=201)
def create_invoice_from_document(
    document_id: str,
    payload: ManualInvoiceCreate = ManualInvoiceCreate(),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.CLIENT:
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="Seuls les administrateurs et comptables peuvent créer des factures.",
        )

    doc = document_service.get_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable.")

    # Verify company access
    if doc.client_id:
        client = client_service.get_client(db, doc.client_id)
        if not client or client.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Accès refusé.")

    if doc.doc_type == "audio":
        raise HTTPException(status_code=400, detail="Impossible de créer une facture depuis une note vocale.")

    # Return existing invoice if one exists
    if doc.invoice:
        return invoice_service._to_read(doc.invoice)

    # Try auto-extraction via Textract
    extracted = {}
    try:
        extracted = textract_service.extract_from_s3(doc.s3_key)
    except Exception:
        pass  # extraction is best-effort

    # Merge: explicit payload fields override extracted values
    invoice = Invoice(
        document_id=doc.id,
        invoice_number=payload.invoice_number or extracted.get("invoice_number") or "",
        supplier_name=payload.supplier_name or extracted.get("supplier_name") or "",
        date=payload.date or extracted.get("date") or "",
        total_amount=payload.total_amount or extracted.get("total_amount") or 0,
        vat_amount=payload.vat_amount or extracted.get("vat_amount") or 0,
        status="to_review",
        direction=payload.direction,
    )
    db.add(invoice)
    db.commit()
    db.refresh(invoice)
    return invoice_service._to_read(invoice)


# ─── Document data extraction (AWS Textract) ────────────────────────────────

@router.post("/documents/{document_id}/extract", response_model=ExtractionResult)
def extract_document_data(
    document_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.CLIENT:
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="Accès refusé.",
        )

    doc = document_service.get_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable.")

    if doc.doc_type == "audio":
        raise HTTPException(status_code=400, detail="L'extraction n'est pas disponible pour les notes vocales.")

    if doc.client_id:
        client = client_service.get_client(db, doc.client_id)
        if not client or client.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Accès refusé.")

    ext = doc.file_name.lower().rsplit(".", 1)[-1] if "." in doc.file_name else ""

    if ext in ("xlsx", "xls"):
        # Excel: use openpyxl-based extraction
        try:
            from app.utils.invoice_parser import InvoiceParser
            pdf_bytes = s3_service.download_file(doc.s3_key)
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(pdf_bytes), read_only=True, data_only=True)
            all_values = []
            for row in wb[wb.sheetnames[0]].iter_rows(max_row=50, values_only=True):
                for cell in row:
                    if cell is not None:
                        all_values.append(str(cell))
            full_text = " ".join(all_values)
            result = InvoiceParser().parse_text(full_text)
            result["confidence"] = min(result.get("confidence", 0), 0.5)
            return ExtractionResult(**result)
        except Exception as exc:
            raise HTTPException(status_code=http_status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Extraction Excel échouée: {exc}")

    if ext not in ("pdf", "jpg", "jpeg", "png"):
        raise HTTPException(
            status_code=http_status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Format non supporté. Utilisez PDF, JPEG, PNG ou XLSX.",
        )

    try:
        result = textract_service.extract_from_s3(doc.s3_key)
    except TextractError as exc:
        raise HTTPException(
            status_code=http_status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        )

    return ExtractionResult(**result)
